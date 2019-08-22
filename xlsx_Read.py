import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

#Getting sheets from the workbook

print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

mySheet = wb.create_sheet('mySheet')
print(wb.sheetnames)

#sheet3 = wb.get_sheet_by_name('Sheet3')
#sheet4 = wb['mySheet']

#Getting cells from the sheets
#active sheet
ws = wb.active
print(ws)
print(ws['A1'])
print(ws['A1'].value)

c= ws['B1']
print('Row {},Column {} is {}'.format(c.row,c.column,c.value))
print('Cell {} is {}\n'.format(c.coordinate,c.value))

print(ws.cell(row =1,column = 2))
print(ws.cell(row =1,column = 2).value)
for i in range(1, 8, 1):
    print(i,ws.cell(row = i,column = 2).value)


colB = ws['B']
row6 = ws[6]
col_range = ws['A:B']
row_range = ws[2:6]
print(colB[1].value)

for col in col_range:
    for cell in col:
        print(cell.value)

for row in ws.iter_rows(min_row = 1,max_row = 2, max_col = 2):
    for cell in row:
        print(cell.value)

print(tuple(ws.rows))

cell_range = ws['A1:B6']

for rowOfCellObjects in cell_range:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('-------------End of Rows-------------')

print('{}*{}'.format(ws.max_row, ws.max_column))


from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(2), get_column_letter(47))
print(column_index_from_string('AAH'))


